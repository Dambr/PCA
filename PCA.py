from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
import os
import numpy as np
import random
import math
from openpyxl import *
from openpyxl.styles import Alignment
class Window():
	def __init__(self):
		self.label = Label( root,
				font = ( 'Tahoma', 14),
				fg = '#2bb928',
				text = 'Точность анализа'
			)
		self.enter= Entry( root, 
				font = ('Tahoma', 14), 
				justify = CENTER,
				fg = '#2bb928',
				borderwidth = '2',
				relief = 'solid',
				width = 8
			)
		self.openExel = Button(root,
			font = ('Tahoma', 12),
			fg = '#0a8218',
			bg = '#cab4ba',
			text = 'Открыть файл Exel'
			)
		self.openExel.bind('<ButtonRelease-1>', self.main)
		self.openExel.bind('<space>', self.main)
		self.label.place(x = 120, y = 50)
		self.enter.place(x = 295, y = 50)
		self.openExel.place(x = 180, y = 150)
	def Centrir(self, X):
		X = np.array(X).T
		sr = []
		for i in range(len(X)):
			sr.append(sum(X[i]) / len(X[i]))
		X = X.T
		for i in range(len(X)):
			for j in range(len(X[i])):
				X[i][j] -= sr[j]
		return X
	def MainComponetnts(self, X, accuracy):
		X_r = len(X)
		X_c = len(X[0])
		P = []
		T = []
		pc = min(X_c, X_r)
		for k in range(pc):
			X = np.matrix(X)
			P1 = np.matrix([[random.random() for i in range(1)] for j in range(X_c)])
			T1 = X * P1
			d0 = T1.T * T1
			P1 = (T1.T * X / (T1.T * T1)).T
			P1 = P1/np.linalg.norm(P1)
			T1 = X * P1
			d = T1.T * T1
			while d - d0 > accuracy:
				P1 = (T1.T * X / (T1.T * T1)).T
				P1 = P1 / np.linalg.norm(P1)
				T1 = X * P1
				d0 = T1.T * T1
				P1 = (T1.T * X / (T1.T * T1)).T
				P1 = P1 / np.linalg.norm(P1)
				T1 = X * P1
				d = T1.T * T1
			X = X - T1 * P1.T
			T.append(T1)
			P.append(P1.T)
		return [T, P, X]
	def main(self, event):
		try:
			accuracy = float(self.enter.get())
		except:
			messagebox.showerror('Ошибка', 'Внесены данные, которые не удается обработать.')
			return
		wbName = filedialog.askopenfilename(initialdir = "/",title = "Открыть Exel файл",filetypes = (("exel files","*.xlsx"),("all files","*.*")))
		try:
			wb = load_workbook(wbName)
		except:
			return
		names = []
		X = []
		for i in range(2, wb.worksheets[0].max_column + 1):
			names.append(wb.worksheets[0].cell(row=1, column=i).value)
		for i in range(2, wb.worksheets[0].max_column + 1):
			X.append([])
			for j in range(2, wb.worksheets[0].max_row + 1):
				X[len(X) - 1].append(wb.worksheets[0].cell(row=j, column=i).value)
		X = np.array(X).T
		X = self.Centrir(X)
		answer = self.MainComponetnts(X, accuracy)
		for i in range(1, len(answer[0])):
			answer[0][0] = np.concatenate((answer[0][0], answer[0][i]), axis = 1)
		for i in range(1, len(answer[1])):
			answer[1][0] = np.concatenate((answer[1][0], answer[1][i]), axis = 0)
		T = np.array(answer[0][0])
		P = np.array(answer[1][0])
		E = np.array(answer[2])
		wb.create_sheet('Центир. вх. матр')
		wb.get_sheet_by_name('Центир. вх. матр').column_dimensions['A'].width = 10.67
		index = 0
		for i in range(2, len(X) + 2):
			wb.get_sheet_by_name('Центир. вх. матр').cell(row=i, column=1).value = 'Значение ' + str(index + 1)
			index += 1
		index = 0
		for i in range(2, len(X[0]) + 2):
			wb.get_sheet_by_name('Центир. вх. матр').cell(row=1, column=i).value = names[index]
			index += 1
		for i in range(2, len(X) + 2):
			for j in range(2, len(X[i - 2]) + 2):
				wb.get_sheet_by_name('Центир. вх. матр').cell(row=i, column=j).value = X[i - 2][j - 2]
		for i in range(1, wb.get_sheet_by_name('Центир. вх. матр').max_column + 1):
			for j in range(1, wb.get_sheet_by_name('Центир. вх. матр').max_row + 1):
				wb.get_sheet_by_name('Центир. вх. матр').cell(row=j, column=i).alignment = Alignment(horizontal='center')
		wb.create_sheet('Матрица счетов')
		wb.get_sheet_by_name('Матрица счетов').column_dimensions['A'].width = 10.67
		index = 0
		for i in range(2, len(X) + 2):
			wb.get_sheet_by_name('Матрица счетов').cell(row=i, column=1).value = 'Значение ' + str(index + 1)
			index += 1
		index = 0
		for i in range(2, len(X[0]) + 2):
			wb.get_sheet_by_name('Матрица счетов').cell(row=1, column=i).value = 't ' + str(index + 1)
			index += 1
		for i in range(2, len(T) + 2):
			for j in range(2, len(T[i - 2]) + 2):
				wb.get_sheet_by_name('Матрица счетов').cell(row=i, column=j).value = round(T[i - 2][j - 2], len(str(accuracy)) - 1)
		for i in range(1, wb.get_sheet_by_name('Матрица счетов').max_column + 1):
			for j in range(1, wb.get_sheet_by_name('Матрица счетов').max_row + 1):
				wb.get_sheet_by_name('Матрица счетов').cell(row=j, column=i).alignment = Alignment(horizontal='center')
		wb.create_sheet('Матрица нагрузок')
		wb.get_sheet_by_name('Матрица нагрузок').column_dimensions['A'].width = 10.67
		index = 0
		for i in range(2, len(X[0]) + 2):
			wb.get_sheet_by_name('Матрица нагрузок').cell(row=i, column=1).value = names[index]
			index += 1
		index = 0
		for i in range(2, len(X[0]) + 2):
			wb.get_sheet_by_name('Матрица нагрузок').cell(row=1, column=i).value = 't ' + str(index + 1)
			index += 1
		for i in range(2, len(P) + 2):
			for j in range(2, len(P[i - 2]) + 2):
				wb.get_sheet_by_name('Матрица нагрузок').cell(row=i, column=j).value = round(P[i - 2][j - 2], len(str(accuracy)) - 1)
		for i in range(1, wb.get_sheet_by_name('Матрица нагрузок').max_column + 1):
			for j in range(1, wb.get_sheet_by_name('Матрица нагрузок').max_row + 1):
				wb.get_sheet_by_name('Матрица нагрузок').cell(row=j, column=i).alignment = Alignment(horizontal='center')
		wb.create_sheet('Матрица остатков')
		wb.get_sheet_by_name('Матрица остатков').column_dimensions['A'].width = 10.67
		index = 0
		for i in range(2, len(X) + 2):
			wb.get_sheet_by_name('Матрица остатков').cell(row=i, column=1).value = 'Значение ' + str(index + 1)
			index += 1
		index = 0
		for i in range(2, len(X[0]) + 2):
			wb.get_sheet_by_name('Матрица остатков').cell(row=1, column=i).value = names[index]
			index += 1
		for i in range(2, len(X) + 2):
			for j in range(2, len(X[i - 2]) + 2):
				wb.get_sheet_by_name('Матрица остатков').cell(row=i, column=j).value = round(E[i - 2][j - 2], len(str(accuracy)) - 1)
		for i in range(1, wb.get_sheet_by_name('Матрица остатков').max_column + 1):
			for j in range(1, wb.get_sheet_by_name('Матрица остатков').max_row + 1):
				wb.get_sheet_by_name('Матрица остатков').cell(row=j, column=i).alignment = Alignment(horizontal='center')
		wb.save(wbName)
		os.system(wbName)
root = Tk()
margin_x = (root.winfo_screenwidth() - root.winfo_reqwidth())
margin_y = (root.winfo_screenheight() - root.winfo_reqheight())
root.title('Факторный анализ')
root.resizable(width=False, height=False)
root.wm_geometry('560x270+%d+%d' % (margin_x/3, margin_y/3))
window = Window()
root.mainloop()